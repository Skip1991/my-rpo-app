from flask import Flask, request, render_template, redirect, url_for, session, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_socketio import SocketIO, emit
from flask_migrate import Migrate
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta
import os
import pandas as pd
from io import BytesIO
import json
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.secret_key = os.urandom(24)
app.config['SQLALCHEMY_DATABASE_URI'] = 'postgresql://postgres:mypassword@localhost/rpo_db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)
socketio = SocketIO(app, cors_allowed_origins="*", async_mode='eventlet')
migrate = Migrate(app, db)

# Определяем get_work_dates в глобальной области видимости
def get_work_dates(start_date, end_date):
    if isinstance(start_date, str):
        start_date = datetime.strptime(start_date, '%Y-%m-%d')
    if isinstance(end_date, str):
        end_date = datetime.strptime(end_date, '%Y-%m-%d')
    dates = []
    current_date = start_date
    while current_date <= end_date:
        dates.append(current_date)
        current_date += timedelta(days=1)
    return dates

# Определяем is_date_in_week в глобальной области видимости
def is_date_in_week(week_date, work_date):
    if not work_date:
        return False
    if isinstance(work_date, str):
        work_date = datetime.strptime(work_date, '%Y-%m-%d')
    if isinstance(week_date, str):
        week_date = datetime.strptime(week_date, '%Y-%m-%d')
    return week_date.strftime('%Y-%m-%d') == work_date.strftime('%Y-%m-%d')

# Модель пользователя
class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(50), unique=True, nullable=False)
    password = db.Column(db.String(128), nullable=False)
    role = db.Column(db.String(20), nullable=False)

# Модель работы (РPO)
class Work(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    organization = db.Column(db.String(100), nullable=False)
    sp = db.Column(db.String(100), nullable=False)
    workshop = db.Column(db.String(100), nullable=False)
    object = db.Column(db.String(100), nullable=False)
    resp_preparation = db.Column(db.String(100), nullable=False)
    resp_execution = db.Column(db.String(100), nullable=False)
    description = db.Column(db.Text, nullable=False)
    work_name = db.Column(db.String(100), nullable=False)
    rpo_type = db.Column(db.String(50), nullable=False)
    approval_adjacent = db.Column(db.String(10), nullable=False)
    risk_level = db.Column(db.String(50), nullable=False)
    working_group = db.Column(db.String(20), nullable=False)
    work_type = db.Column(db.String(20), nullable=False)
    start_date = db.Column(db.Date, nullable=False)
    end_date = db.Column(db.Date, nullable=False)
    pbotos_approved = db.Column(db.Boolean, default=False)
    cits_approved = db.Column(db.Boolean, default=False)
    gi_approved = db.Column(db.Boolean, default=False)
    sp_approved = db.Column(db.Boolean, default=False)
    rejected_by = db.Column(db.String(50), nullable=True)
    rejection_reason = db.Column(db.Text, nullable=True)
    approval_status = db.Column(db.String(50), default='В процессе')
    execution_status = db.Column(db.String(50), default='В процессе выполнения')
    work_times = db.Column(db.JSON, default={})
    comments = db.Column(db.Text, nullable=True)
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    nd_status = db.Column(db.String(50), default='Планируется')

# Модель уведомлений
class Notification(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    message = db.Column(db.Text, nullable=False)
    timestamp = db.Column(db.DateTime, default=datetime.utcnow)
    action_url = db.Column(db.String(100), nullable=True)
    is_active = db.Column(db.Boolean, default=True)

# Функция для генерации дней недели и форматирования дат
@app.context_processor
def utility_processor():
    def get_week_dates(start_date=None, end_date=None):
        if not start_date or not end_date:
            today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
            start = today - timedelta(days=today.weekday())
            end = start + timedelta(days=6)
        else:
            try:
                start = datetime.strptime(start_date, '%Y-%m-%d') if isinstance(start_date, str) else start_date
                start = start.replace(hour=0, minute=0, second=0, microsecond=0) - timedelta(days=start.weekday())
                end = datetime.strptime(end_date, '%Y-%m-%d') if isinstance(end_date, str) else end_date
            except ValueError:
                today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
                start = today - timedelta(days=today.weekday())
                end = start + timedelta(days=6)
        days = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота', 'Воскресенье']
        week = []
        current_date = start
        while current_date <= end:
            day_idx = (current_date.weekday()) % 7
            week.append((days[day_idx][:2], current_date.strftime('%d.%m.%Y'), current_date.strftime('%Y-%m-%d')))
            current_date += timedelta(days=1)
        return week
    
    def format_date(date):
        if isinstance(date, str):
            date = datetime.strptime(date, '%Y-%m-%d')
        return date.strftime('%d.%m.%Y')
    
    unread_notifications = Notification.query.filter_by(user_id=session.get('user_id'), is_active=True).count() if session.get('user_id') else 0
    return dict(get_week_dates=get_week_dates, format_date=format_date, is_date_in_week=is_date_in_week, get_work_dates=get_work_dates, unread_notifications=unread_notifications)

# Инициализация базы данных и создание пользователей
with app.app_context():
    db.create_all()
    if not User.query.first():
        users = [
            ('admin', 'admin123', 'admin'),
            ('user', 'user123', 'user'),
            ('pbotos', 'pbotos123', 'pbotos'),
            ('cits', 'cits123', 'cits'),
            ('gi', 'gi123', 'gi'),
            ('sp_user', 'sp123', 'sp')
        ]
        for username, password, role in users:
            db.session.add(User(username=username, password=generate_password_hash(password), role=role))
        db.session.commit()

# Логин
@app.route('/', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        user = User.query.filter_by(username=username).first()
        if user and check_password_hash(user.password, password):
            session['user_id'] = user.id
            session['username'] = user.username
            session['role'] = user.role
            return redirect(url_for('main'))
        return render_template('login.html', error='Неверный логин или пароль')
    return render_template('login.html')

# Главная страница
@app.route('/main')
def main():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    return render_template('main.html', username=session['username'], role=session['role'])

# Уведомления
@app.route('/notifications', methods=['GET', 'POST'])
def notifications():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    if request.method == 'POST':
        if 'delete_notification' in request.form:
            notification_id = request.form['notification_id']
            notification = Notification.query.get(notification_id)
            if notification and notification.user_id == session['user_id']:
                db.session.delete(notification)
                db.session.commit()
        elif 'mark_viewed' in request.form:
            notification_id = request.form['notification_id']
            notification = Notification.query.get(notification_id)
            if notification and notification.user_id == session['user_id']:
                notification.is_active = False
                db.session.commit()
        return redirect(url_for('notifications'))
    
    notifications = Notification.query.filter_by(user_id=session['user_id']).order_by(Notification.timestamp.desc()).all()
    return render_template('notifications.html', notifications=notifications)

# Внесение РPO
@app.route('/rpo_entry', methods=['GET', 'POST'])
def rpo_entry():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    if request.method == 'POST':
        work = Work(
            organization=request.form['organization'],
            sp=request.form['sp'],
            workshop=request.form['workshop'],
            object=request.form['object'],
            resp_preparation=request.form['resp_preparation'],
            resp_execution=request.form['resp_execution'],
            description=request.form['description'],
            work_name=request.form['work_name'],
            rpo_type=request.form['rpo_type'],
            approval_adjacent=request.form['approval_adjacent'],
            risk_level=request.form['risk_level'],
            working_group=request.form['working_group'],
            work_type=request.form['work_type'],
            start_date=request.form['start_date'],
            end_date=request.form['end_date'],
            created_by=session['user_id'],
            nd_status='Планируется'
        )
        db.session.add(work)
        db.session.commit()
        
        approving_roles = ['pbotos', 'cits', 'gi', 'sp']
        approving_users = User.query.filter(User.role.in_(approving_roles)).all()
        if work.work_type in ['Не плановые', 'Аварийные']:
            for user in approving_users:
                notification = Notification(
                    user_id=user.id,
                    message=f"Добавлена новая работа '{work.work_name}' ({work.work_type}).",
                    action_url=url_for('rpo_execution')
                )
                db.session.add(notification)
                socketio.emit('notification', {
                    'user_id': user.id,
                    'message': f"Добавлена новая работа '{work.work_name}' ({work.work_type}).",
                    'action_url': url_for('rpo_execution')
                }, namespace='/')
        else:
            for user in approving_users:
                notification = Notification(
                    user_id=user.id,
                    message=f"Новая работа '{work.work_name}' поступила на согласование.",
                    action_url=url_for('rpo_approval')
                )
                db.session.add(notification)
                socketio.emit('notification', {
                    'user_id': user.id,
                    'message': f"Новая работа '{work.work_name}' поступила на согласование.",
                    'action_url': url_for('rpo_approval')
                }, namespace='/')
        db.session.commit()
        
        socketio.emit('update_works', {'message': 'Новая работа добавлена', 'work_id': work.id}, namespace='/')
        return redirect(url_for('rpo_entry'))
    return render_template('rpo_entry.html')

# Согласование РPO
@app.route('/rpo_approval', methods=['GET', 'POST'])
def rpo_approval():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    if request.method == 'POST':
        if 'filter' in request.form:
            start_date = request.form['start_date']
            end_date = request.form['end_date']
            session['filter_start_date'] = start_date
            session['filter_end_date'] = end_date
            works = Work.query.filter(
                Work.start_date >= start_date,
                Work.end_date <= end_date,
                Work.work_type == 'Плановые'
            ).all()
            return render_template('rpo_approval.html', works=works, start_date=start_date, end_date=end_date)
        
        if 'action' not in request.form:
            print("Ошибка: 'action' отсутствует в запросе")
            return redirect(url_for('rpo_approval'))
        
        work_id = request.form['work_id']
        action = request.form['action']
        work = db.session.get(Work, work_id)
        role = session['role']
        print(f"Нажата кнопка: action={action}, work_id={work_id}, role={role}")
        
        if action == 'approve' and role in ['pbotos', 'cits', 'gi', 'sp']:
            if role == 'pbotos':
                work.pbotos_approved = True
                print("Согласовано ПБОТОС")
            elif role == 'cits':
                work.cits_approved = True
                print("Согласовано ЦИТС")
            elif role == 'gi':
                work.gi_approved = True
                print("Согласовано ГИ")
            elif role == 'sp':
                work.sp_approved = True
                print("Согласовано СП")
            notifications = Notification.query.filter_by(user_id=session['user_id'], message=f"Новая работа '{work.work_name}' поступила на согласование.").all()
            for notification in notifications:
                notification.is_active = False
            if work.pbotos_approved and work.cits_approved and work.gi_approved and work.sp_approved:
                work.approval_status = 'Согласовано'
                work.rejected_by = None
                notification = Notification(
                    user_id=work.created_by,
                    message=f"Работа '{work.work_name}' согласована всеми."
                )
                db.session.add(notification)
                socketio.emit('notification', {
                    'user_id': work.created_by,
                    'message': f"Работа '{work.work_name}' согласована всеми."
                }, namespace='/')
                print("Все согласовано")
        elif action == 'reject':
            work.rejected_by = role
            reason = request.form.get('reason', 'Не указана причина')
            work.rejection_reason = reason
            notification = Notification(
                user_id=work.created_by,
                message=f"Работа '{work.work_name}' отклонена пользователем {role}. Причина: {reason}"
            )
            db.session.add(notification)
            socketio.emit('notification', {
                'user_id': work.created_by,
                'message': f"Работа '{work.work_name}' отклонена пользователем {role}. Причина: {reason}"
            }, namespace='/')
            notifications = Notification.query.filter_by(user_id=session['user_id'], message=f"Новая работа '{work.work_name}' поступила на согласование.").all()
            for notification in notifications:
                notification.is_active = False
            db.session.delete(work)
            print(f"Работа удалена: {role}, причина: {reason}")
        
        db.session.commit()
        socketio.emit('update_works', {
            'work_id': work_id,
            'approval_status': work.approval_status,
            'pbotos_approved': work.pbotos_approved,
            'cits_approved': work.cits_approved,
            'gi_approved': work.gi_approved,
            'sp_approved': work.sp_approved,
            'rejected_by': work.rejected_by,
            'work_times': work.work_times,
            'nd_status': work.nd_status
        }, namespace='/')
        print(f"Событие отправлено через WebSocket: {json.dumps({'work_id': work_id, 'approval_status': work.approval_status, 'pbotos_approved': work.pbotos_approved, 'cits_approved': work.cits_approved, 'gi_approved': work.gi_approved, 'sp_approved': work.sp_approved, 'rejected_by': work.rejected_by, 'work_times': work.work_times, 'nd_status': work.nd_status}, ensure_ascii=False)}")
        return redirect(url_for('rpo_approval'))
    works = Work.query.filter(Work.work_type == 'Плановые').all()
    start_date = session.get('filter_start_date', datetime.now().strftime('%Y-%m-%d'))
    end_date = session.get('filter_end_date', (datetime.now() + timedelta(days=7)).strftime('%Y-%m-%d'))
    return render_template('rpo_approval.html', works=works, start_date=start_date, end_date=end_date)

# Фактическое выполнение РPO с текстовыми полями и сохранением структуры таблицы
@app.route('/rpo_execution', methods=['GET', 'POST'])
def rpo_execution():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    if request.method == 'POST':
        if 'filter' in request.form:
            start_date = request.form['start_date']
            end_date = request.form['end_date']
            session['filter_start_date'] = start_date
            session['filter_end_date'] = end_date
            works = Work.query.filter(
                ((Work.approval_status == 'Согласовано') | (Work.work_type.in_(['Не плановые', 'Аварийные']))) &
                (Work.start_date >= start_date) & (Work.end_date <= end_date)
            ).all()
            return render_template('rpo_execution.html', works=works, start_date=start_date, end_date=end_date)
        elif 'save_status' in request.form:
            work_id = request.form.get('work_id')
            date = request.form.get('date')
            periods = ['утро', 'день', 'вечер', 'ночь']  # Используем русские названия для соответствия шаблону
            work = db.session.get(Work, work_id)  # Используем Session.get()
            if not work:
                print(f"Работа с ID {work_id} не найдена")
                return redirect(url_for('rpo_execution'))

            if not work.work_times:
                work.work_times = {}
            if date not in work.work_times:
                work.work_times[date] = {}

            # Обработка статусов времени суток
            for period in periods:
                status_key = f"status_{work_id}_{date}_{period.lower()}"
                status = request.form.get(status_key)
                if status:
                    # Валидация статусов времени суток
                    valid_statuses = ['Подготовка', 'Проведение', 'Завершено', 'Срыв']
                    status_str = status.encode('utf-8').decode('utf-8') if status else ''
                    if status_str in valid_statuses:
                        work.work_times[date][period.lower()] = status_str
                    else:
                        work.work_times[date][period.lower()] = ''  # Оставляем пустым, если некорректный статус

            print(f"Обновлён work_times: {json.dumps(work.work_times, ensure_ascii=False)}")

            # Обработка статуса НД
            nd_status = request.form.get('nd_status')
            if nd_status:
                valid_nd_statuses = ['Оформление', 'На согласовании', 'Утверждение', 'Выдан', 'Не требуется', 'Планируется']
                nd_status_str = nd_status.encode('utf-8').decode('utf-8')
                if nd_status_str in valid_nd_statuses:
                    work.nd_status = nd_status_str
                    print(f"Обновлён nd_status: {nd_status_str}")
                else:
                    print(f"Некорректный nd_status: {nd_status_str}, установлен по умолчанию 'Планируется'")
                    work.nd_status = 'Планируется'
            else:
                print(f"nd_status не передан, оставлен текущий: {work.nd_status or 'Планируется'}")

            # Обновляем execution_status на основе статуса выполнения
            work_dates = [d for d in work.work_times.keys() if isinstance(d, str) and d]
            if work_dates:
                if all(work.work_times.get(d, {}).get(p, '') == 'Завершено' for d in work_dates for p in periods if p in work.work_times.get(d, {})):
                    work.execution_status = 'Выполнено'
                elif any(work.work_times.get(d, {}).get(p, '') == 'Срыв' for d in work_dates for p in periods if p in work.work_times.get(d, {})):
                    work.execution_status = 'Не выполнено'
                else:
                    work.execution_status = 'В процессе выполнения'
            else:
                work.execution_status = 'В процессе выполнения'

            print(f"Проверка перед сохранением: work_times={json.dumps(work.work_times, ensure_ascii=False)}, execution_status={work.execution_status}")

            try:
                db.session.commit()
                print(f"Сохранены изменения для work_id={work_id}")
            except Exception as e:
                db.session.rollback()
                print(f"Ошибка при сохранении: {str(e)}")
                return redirect(url_for('rpo_execution'))

            # Перечитываем данные для обновления
            work = db.session.get(Work, work_id)
            update_data = {
                'work_id': work_id,
                'execution_status': work.execution_status,
                'work_times': work.work_times,
                'nd_status': work.nd_status
            }
            socketio.emit('update_works', update_data, namespace='/')
            print(f"Событие update_works отправлено: {json.dumps(update_data, ensure_ascii=False)}")

            # Сохраняем текущие фильтры и возвращаем шаблон без сброса
            start_date = session.get('filter_start_date', request.form.get('start_date', datetime.now().strftime('%Y-%m-%d')))
            end_date = session.get('filter_end_date', request.form.get('end_date', (datetime.now() + timedelta(days=7)).strftime('%Y-%m-%d')))
            works = Work.query.filter(
                ((Work.approval_status == 'Согласовано') | (Work.work_type.in_(['Не плановые', 'Аварийные']))) &
                (Work.start_date >= start_date) & (Work.end_date <= end_date)
            ).all()
            return render_template('rpo_execution.html', works=works, start_date=start_date, end_date=end_date)
        elif 'comments' in request.form:
            work_id = request.form.get('work_id')
            comments = request.form.get('comments', '')
            if work_id:
                work = db.session.get(Work, work_id)  # Используем Session.get()
                if not work:
                    print(f"Работа с ID {work_id} не найдена")
                    return redirect(url_for('rpo_execution'))
                
                work.comments = comments
                try:
                    db.session.commit()
                    print(f"Сохранён комментарий для work_id={work_id}: {comments}")
                except Exception as e:
                    db.session.rollback()
                    print(f"Ошибка при сохранении комментария: {str(e)}")
                    return redirect(url_for('rpo_execution'))

                # Перечитываем данные для обновления
                work = db.session.get(Work, work_id)  # Перечитываем для актуальных данных
                update_data = {
                    'work_id': work_id,
                    'execution_status': work.execution_status,
                    'work_times': work.work_times,
                    'nd_status': work.nd_status
                }
                socketio.emit('update_works', update_data, namespace='/')
                print(f"Событие update_works отправлено: {json.dumps(update_data, ensure_ascii=False)}")
            start_date = session.get('filter_start_date', request.form.get('start_date', datetime.now().strftime('%Y-%m-%d')))
            end_date = session.get('filter_end_date', request.form.get('end_date', (datetime.now() + timedelta(days=7)).strftime('%Y-%m-%d')))
            works = Work.query.filter(
                ((Work.approval_status == 'Согласовано') | (Work.work_type.in_(['Не плановые', 'Аварийные']))) &
                (Work.start_date >= start_date) & (Work.end_date <= end_date)
            ).all()
            return render_template('rpo_execution.html', works=works, start_date=start_date, end_date=end_date)
        else:
            return redirect(url_for('rpo_execution'))
    
    works = Work.query.filter((Work.approval_status == 'Согласовано') | (Work.work_type.in_(['Не плановые', 'Аварийные']))).all()
    start_date = session.get('filter_start_date', datetime.now().strftime('%Y-%m-%d'))
    end_date = session.get('filter_end_date', (datetime.now() + timedelta(days=7)).strftime('%Y-%m-%d'))
    return render_template('rpo_execution.html', works=works, start_date=start_date, end_date=end_date)

# Статистика
@app.route('/statistics', methods=['GET', 'POST'])
def statistics():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    if request.method == 'POST':
        start_date = request.form['start_date']
        end_date = request.form['end_date']
        stats = compute_statistics(start_date, end_date)
        return render_template('statistics.html', stats=stats, start_date=start_date, end_date=end_date)
    stats = compute_statistics()
    start_date = datetime.now().strftime('%Y-%m-%d')
    end_date = (datetime.now() + timedelta(days=7)).strftime('%Y-%m-%d')
    return render_template('statistics.html', stats=stats, start_date=start_date, end_date=end_date)

def compute_statistics(start_date=None, end_date=None):
    query = Work.query
    if start_date and end_date:
        query = query.filter(Work.start_date >= start_date, Work.end_date <= end_date)
    stats = {
        'planned': query.filter_by(work_type='Плановые', approval_status='Согласовано').count(),
        'unplanned': query.filter(Work.work_type.in_(['Не плановые', 'Аварийные'])).count(),
        'completed_planned': query.filter_by(work_type='Плановые', approval_status='Согласовано', execution_status='Выполнено').count(),
        'uncompleted_planned': query.filter_by(work_type='Плановые', approval_status='Согласовано', execution_status='Не выполнено').count(),
        'in_progress': query.filter_by(work_type='Плановые', approval_status='Согласовано', execution_status='В процессе выполнения').count()
    }
    stats['total'] = stats['planned'] + stats['unplanned']
    return stats

# Архив
@app.route('/archive', methods=['GET', 'POST'])
def archive():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    if request.method == 'POST':
        if 'filter' in request.form:
            start_date = request.form['start_date']
            end_date = request.form['end_date']
            session['filter_start_date'] = start_date
            session['filter_end_date'] = end_date
            
            if isinstance(start_date, str):
                start_date = datetime.strptime(start_date, '%Y-%m-%d')
            if isinstance(end_date, str):
                end_date = datetime.strptime(end_date, '%Y-%m-%d')
            
            approval_works = Work.query.filter(
                Work.start_date >= start_date,
                Work.end_date <= end_date,
                Work.work_type == 'Плановые',
                Work.approval_status == 'Согласовано'
            ).all()
            
            for work in approval_works:
                if isinstance(work.start_date, str):
                    work.start_date = datetime.strptime(work.start_date, '%Y-%m-%d')
                if isinstance(work.end_date, str):
                    work.end_date = datetime.strptime(work.end_date, '%Y-%m-%d')
                work.dates = get_work_dates(work.start_date, work.end_date)
                print(f"Work ID: {work.id}, Work Type: {work.work_type}, Start Date: {work.start_date}, End Date: {work.end_date}, Dates: {[d.strftime('%Y-%m-%d') for d in work.dates]}, Approval Status: {work.approval_status}")
            
            execution_works = Work.query.filter(
                ((Work.approval_status == 'Согласовано') | (Work.work_type.in_(['Не плановые', 'Аварийные']))) &
                (Work.start_date >= start_date) & (Work.end_date <= end_date)
            ).all()
            
            for work in execution_works:
                if isinstance(work.start_date, str):
                    work.start_date = datetime.strptime(work.start_date, '%Y-%m-%d')
                if isinstance(work.end_date, str):
                    work.end_date = datetime.strptime(work.end_date, '%Y-%m-%d')
                if work.work_type == 'Плановые' and work.approval_status == 'Согласовано':
                    work.dates = get_work_dates(work.start_date, work.end_date)
                else:
                    work.dates = []
            
            works = {}
            for work in approval_works + execution_works:
                if work.id not in works:
                    works[work.id] = work
            works = list(works.values())
            
            return render_template('archive.html', works=works, start_date=start_date.strftime('%Y-%m-%d'), end_date=end_date.strftime('%Y-%m-%d'))
        elif 'export' in request.form:
            start_date = session.get('filter_start_date', datetime.now().strftime('%Y-%m-%d'))
            end_date = session.get('filter_end_date', (datetime.now() + timedelta(days=7)).strftime('%Y-%m-%d'))
            
            if isinstance(start_date, str):
                start_date = datetime.strptime(start_date, '%Y-%m-%d')
            if isinstance(end_date, str):
                end_date = datetime.strptime(end_date, '%Y-%m-%d')
            
            approval_works = Work.query.filter(
                Work.start_date >= start_date,
                Work.end_date <= end_date,
                Work.work_type == 'Плановые',
                Work.approval_status == 'Согласовано'
            ).all()
            
            for work in approval_works:
                if isinstance(work.start_date, str):
                    work.start_date = datetime.strptime(work.start_date, '%Y-%m-%d')
                if isinstance(work.end_date, str):
                    work.end_date = datetime.strptime(work.end_date, '%Y-%m-%d')
                print(f"Export Work ID: {work.id}, Work Type: {work.work_type}, Start Date: {work.start_date}, End Date: {work.end_date}, Dates: {[d.strftime('%Y-%m-%d') for d in get_work_dates(work.start_date, work.end_date)]}, Approval Status: {work.approval_status}")
            
            execution_works = Work.query.filter(
                ((Work.approval_status == 'Согласовано') | (Work.work_type.in_(['Не плановые', 'Аварийные']))) &
                (Work.start_date >= start_date) & (Work.end_date <= end_date)
            ).all()
            
            for work in execution_works:
                if isinstance(work.start_date, str):
                    work.start_date = datetime.strptime(work.start_date, '%Y-%m-%d')
                if isinstance(work.end_date, str):
                    work.end_date = datetime.strptime(work.end_date, '%Y-%m-%d')
                if work.work_type == 'Плановые' and work.approval_status == 'Согласовано':
                    work.dates = get_work_dates(work.start_date, work.end_date)
                else:
                    work.dates = []
            
            works = {}
            for work in approval_works + execution_works:
                if work.id not in works:
                    works[work.id] = work
            works = list(works.values())
            
            template_path = os.path.join(app.static_folder, 'templates', 'template.xlsx')
            if not os.path.exists(template_path):
                return "Шаблон Excel не найден", 404

            wb = load_workbook(template_path)
            ws = wb.active

            start_row = 19  # Данные начинаются с 19-й строки

            week_days = ['Пн', 'Вт', 'Ср', 'Чт', 'Пт', 'Сб', 'Вс']
            dates = [datetime.strptime(start_date.strftime('%Y-%m-%d'), '%Y-%m-%d') + timedelta(days=i) for i in range((end_date - start_date).days + 1) if (start_date + timedelta(days=i)).weekday() < 7][:7]

            for i, work in enumerate(works, 1):
                current_row = start_row + i - 1
                
                ws.cell(row=current_row, column=1, value=i)  # №
                ws.cell(row=current_row, column=2, value=work.organization)  # Организация
                ws.cell(row=current_row, column=3, value=work.sp)  # СП
                ws.cell(row=current_row, column=4, value=work.workshop)  # Цех
                ws.cell(row=current_row, column=5, value=work.object)  # Объект
                ws.cell(row=current_row, column=6, value=work.resp_preparation)  # Ответственный за подготовку
                ws.cell(row=current_row, column=7, value=work.resp_execution)  # Ответственный за проведение
                ws.cell(row=current_row, column=8, value=work.description)  # Описание работ
                ws.cell(row=current_row, column=9, value=work.work_name)  # Наименование работ
                ws.cell(row=current_row, column=10, value=work.rpo_type)  # Вид РPO
                ws.cell(row=current_row, column=11, value=work.approval_adjacent)  # Согласование со смежными СП
                ws.cell(row=current_row, column=12, value=work.risk_level)  # Уровень оценки риска
                ws.cell(row=current_row, column=13, value=work.working_group)  # Выезд рабочей группы

                # Дни недели и даты из "Согласование РPO" (столбцы 14–20, ПЛАНИРОВАНИЕ) — используем весь диапазон start_date и end_date для плановых и согласованных работ
                date_col = 14
                for date in dates:
                    d_str = date.strftime('%Y-%m-%d')
                    if work.work_type == 'Плановые' and work.approval_status == 'Согласовано' and work.dates and any(is_date_in_week(date, d) for d in work.dates):
                        ws.cell(row=current_row, column=date_col, value=f"Плановая: {date.strftime('%d.%m.%Y')}")  # Используем весь диапазон дат
                    else:
                        ws.cell(row=current_row, column=date_col, value='')  # Пустая ячейка, если не плановая или не согласована
                    date_col += 1

                # Примечания, Статус НД, Текущий этап (столбцы 21–23, обычные заголовки)
                ws.cell(row=current_row, column=21, value=work.comments or '')  # Примечания (условия выполнения работ)
                ws.cell(row=current_row, column=22, value=work.nd_status)  # Статус НД
                ws.cell(row=current_row, column=23, value=work.work_type)  # Текущий этап

                # Дни недели, даты и время суток из "Фактическое выполнение РPO" (столбцы 24–51, ВЫПОЛНЕНИЕ) — только время суток, без даты
                day_col = 24
                for date in dates:
                    d_str = date.strftime('%Y-%m-%d')
                    # Убрали столбец с датой, оставили только время суток
                    ws.cell(row=current_row, column=day_col, value=work.work_times.get(d_str, {}).get('утро', '-'))  # Утро — данные из work_times с дефолтом "-"
                    day_col += 1
                    ws.cell(row=current_row, column=day_col, value=work.work_times.get(d_str, {}).get('день', '-'))  # День — данные из work_times с дефолтом "-"
                    day_col += 1
                    ws.cell(row=current_row, column=day_col, value=work.work_times.get(d_str, {}).get('вечер', '-'))  # Вечер — данные из work_times с дефолтом "-"
                    day_col += 1
                    ws.cell(row=current_row, column=day_col, value=work.work_times.get(d_str, {}).get('ночь', '-'))  # Ночь — данные из work_times с дефолтом "-"
                    day_col += 1

                # Оставшиеся столбцы (Плановая, Статус выполнения, Примечание)
                ws.cell(row=current_row, column=52, value=work.work_type)  # Плановая — тип РPO
                ws.cell(row=current_row, column=53, value=work.execution_status)  # Статус выполнения
                ws.cell(row=current_row, column=54, value=work.rejection_reason or work.comments or '')  # Примечание

            output = BytesIO()
            wb.save(output)
            output.seek(0)

            return send_file(output, download_name='weekly_planning.xlsx', as_attachment=True)
        
        work_id = request.form['work_id']
        notes = request.form['notes']
        work = db.session.get(Work, work_id)
        work.comments = notes
        db.session.commit()
        return redirect(url_for('archive'))
    
    start_date = session.get('filter_start_date', datetime.now().strftime('%Y-%m-%d'))
    end_date = session.get('filter_end_date', (datetime.now() + timedelta(days=7)).strftime('%Y-%m-%d'))
    
    if isinstance(start_date, str):
        start_date = datetime.strptime(start_date, '%Y-%m-%d')
    if isinstance(end_date, str):
        end_date = datetime.strptime(end_date, '%Y-%m-%d')
    
    approval_works = Work.query.filter(
        Work.start_date >= start_date,
        Work.end_date <= end_date,
        Work.work_type == 'Плановые',
        Work.approval_status == 'Согласовано'
    ).all()
    
    for work in approval_works:
        if isinstance(work.start_date, str):
            work.start_date = datetime.strptime(work.start_date, '%Y-%m-%d')
        if isinstance(work.end_date, str):
            work.end_date = datetime.strptime(work.end_date, '%Y-%m-%d')
        work.dates = get_work_dates(work.start_date, work.end_date)
        print(f"Work ID: {work.id}, Work Type: {work.work_type}, Start Date: {work.start_date}, End Date: {work.end_date}, Dates: {[d.strftime('%Y-%m-%d') for d in work.dates]}, Approval Status: {work.approval_status}")
    
    execution_works = Work.query.filter(
        ((Work.approval_status == 'Согласовано') | (Work.work_type.in_(['Не плановые', 'Аварийные']))) &
        (Work.start_date >= start_date) & (Work.end_date <= end_date)
    ).all()
    
    for work in execution_works:
        if isinstance(work.start_date, str):
            work.start_date = datetime.strptime(work.start_date, '%Y-%m-%d')
        if isinstance(work.end_date, str):
            work.end_date = datetime.strptime(work.end_date, '%Y-%m-%d')
        if work.work_type == 'Плановые' and work.approval_status == 'Согласовано':
            work.dates = get_work_dates(work.start_date, work.end_date)
        else:
            work.dates = []
    
    works = {}
    for work in approval_works + execution_works:
        if work.id not in works:
            works[work.id] = work
    works = list(works.values())
    
    return render_template('archive.html', works=works, start_date=start_date.strftime('%Y-%m-%d'), end_date=end_date.strftime('%Y-%m-%d'))

# Выход
@app.route('/logout')
def logout():
    session.pop('user_id', None)
    session.pop('filter_start_date', None)
    session.pop('filter_end_date', None)
    return redirect(url_for('login'))

# Количество уведомлений
@app.route('/notifications_count')
def notifications_count():
    if 'user_id' not in session:
        return jsonify({'count': 0})
    count = Notification.query.filter_by(user_id=session['user_id'], is_active=True).count()
    return jsonify({'count': count})

# WebSocket-события
@socketio.on('connect', namespace='/')
def handle_connect():
    print('Клиент подключён к /')

@socketio.on('disconnect', namespace='/')
def handle_disconnect():
    print('Клиент отключён от /')

@socketio.on('reconnect', namespace='/')
def handle_reconnect():
    print('Клиент переподключён к /')

if __name__ == '__main__':
    socketio.run(app, host='0.0.0.0', port=5000, debug=True, allow_unsafe_werkzeug=True)